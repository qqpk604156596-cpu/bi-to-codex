#!/usr/bin/env python3
"""Load UCI Online Retail II into MySQL after local approval and setup."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping


FIELD_ALIASES = {
    "InvoiceNo": ("InvoiceNo", "Invoice"),
    "StockCode": ("StockCode",),
    "Description": ("Description",),
    "Quantity": ("Quantity",),
    "InvoiceDate": ("InvoiceDate",),
    "UnitPrice": ("UnitPrice", "Price"),
    "CustomerID": ("CustomerID", "Customer ID"),
    "Country": ("Country",),
}
PROJECT_ROOT = Path(__file__).resolve().parents[2]
REQUIRED_ENVIRONMENT = (
    "MYSQL_HOST",
    "MYSQL_PORT",
    "MYSQL_DATABASE",
    "MYSQL_USER",
    "MYSQL_PASSWORD",
    "MYSQL_SOURCE_TABLE",
)
IDENTIFIER = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def validate_workbook_path(workbook_path: Path, project_root: Path) -> Path:
    resolved_workbook = workbook_path.resolve()
    resolved_root = project_root.resolve()
    try:
        resolved_workbook.relative_to(resolved_root)
    except ValueError as error:
        raise ValueError("workbook_path_outside_project") from error
    if resolved_workbook.suffix.lower() != ".xlsx":
        raise ValueError("workbook_extension_invalid")
    if not resolved_workbook.is_file():
        raise ValueError("workbook_not_found")
    return resolved_workbook


def source_value(row: Mapping[str, Any], standard_name: str) -> Any:
    for source_name in FIELD_ALIASES[standard_name]:
        if source_name in row:
            return row[source_name]
    return None


def text_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def decimal_value(value: Any, error_code: str) -> Decimal:
    try:
        return Decimal(text_value(value))
    except InvalidOperation as error:
        raise ValueError(error_code) from error


def customer_id_value(value: Any) -> str | None:
    text = text_value(value)
    if not text:
        return None
    try:
        numeric = Decimal(text)
    except InvalidOperation:
        return text
    if numeric == numeric.to_integral_value():
        return str(int(numeric))
    return text


def classify_row(row: Mapping[str, Any]) -> dict[str, Any]:
    """Normalize one UCI source row without deleting cancellations or unknown customers."""

    invoice_no = text_value(source_value(row, "InvoiceNo"))
    stock_code = text_value(source_value(row, "StockCode"))
    description = text_value(source_value(row, "Description")) or None
    country = text_value(source_value(row, "Country"))
    invoice_date = source_value(row, "InvoiceDate")
    if not invoice_no or not stock_code or not country or invoice_date is None:
        raise ValueError("required_source_field_missing")

    quantity = decimal_value(source_value(row, "Quantity"), "quantity_invalid")
    if quantity != quantity.to_integral_value():
        raise ValueError("quantity_not_integral")

    unit_price = decimal_value(source_value(row, "UnitPrice"), "unit_price_invalid")
    if unit_price < 0:
        raise ValueError("unit_price_negative")

    customer_id = customer_id_value(source_value(row, "CustomerID"))
    return {
        "InvoiceNo": invoice_no,
        "StockCode": stock_code,
        "Description": description,
        "Quantity": int(quantity),
        "InvoiceDate": invoice_date,
        "UnitPrice": format(unit_price.quantize(Decimal("0.01")), "f"),
        "CustomerID": customer_id,
        "Country": country,
        "IsCancellation": invoice_no.upper().startswith("C"),
        "UnknownCustomer": customer_id is None,
        "ZeroPrice": unit_price == 0,
    }


def load_dotenv(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def required_environment(env_file: Path) -> dict[str, str]:
    local = load_dotenv(env_file) if env_file.is_file() else {}
    values = {name: local.get(name, os.environ.get(name, "")) for name in REQUIRED_ENVIRONMENT}
    missing = [name for name, value in values.items() if not value]
    if missing:
        raise ValueError("mysql_environment_missing: " + ", ".join(missing))
    if not IDENTIFIER.fullmatch(values["MYSQL_DATABASE"]):
        raise ValueError("mysql_database_identifier_invalid")
    if not IDENTIFIER.fullmatch(values["MYSQL_SOURCE_TABLE"]):
        raise ValueError("mysql_source_identifier_invalid")
    return values


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def ensure_product_refresh_schema(cursor: Any, database: str, table: str) -> None:
    """Idempotently migrate the product description aggregate to its refresh-optimized schema."""
    cursor.execute(
        """
        SELECT DATA_TYPE, CHARACTER_MAXIMUM_LENGTH
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s AND COLUMN_NAME = 'DescriptionText'
        """,
        (database, table),
    )
    column = cursor.fetchone()
    if column is None:
        raise RuntimeError("description_text_column_missing")
    if tuple(column) != ("varchar", 512):
        cursor.execute(
            f"SELECT COUNT(*), MAX(CHAR_LENGTH(`DescriptionText`)) FROM `{database}`.`{table}`"
        )
        rows_before, maximum_length = cursor.fetchone()
        if maximum_length is not None and int(maximum_length) > 512:
            raise ValueError("description_text_exceeds_512")
        cursor.execute(
            f"ALTER TABLE `{database}`.`{table}` MODIFY COLUMN `DescriptionText` VARCHAR(512) NULL"
        )
        cursor.execute(f"SELECT COUNT(*) FROM `{database}`.`{table}`")
        if int(cursor.fetchone()[0]) != int(rows_before):
            raise RuntimeError("refresh_schema_migration_row_count_changed")

    cursor.execute(
        """
        SELECT INDEX_NAME, SEQ_IN_INDEX, COLUMN_NAME
        FROM INFORMATION_SCHEMA.STATISTICS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
        ORDER BY INDEX_NAME, SEQ_IN_INDEX
        """,
        (database, table),
    )
    indexes: dict[str, list[str]] = {}
    for index_name, _sequence, column_name in cursor.fetchall():
        indexes.setdefault(str(index_name), []).append(str(column_name))
    if ["StockCode", "DescriptionText"] not in indexes.values():
        cursor.execute(
            f"ALTER TABLE `{database}`.`{table}` "
            "ADD INDEX `ix_product_cover` (`StockCode`, `DescriptionText`)"
        )


def create_schema(cursor: Any, database: str, table: str, *, replace: bool) -> None:
    cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{database}` CHARACTER SET utf8mb4 COLLATE utf8mb4_0900_ai_ci")
    cursor.execute(f"USE `{database}`")
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS `{table}` (
            SourceRowId BIGINT NOT NULL PRIMARY KEY,
            SourceSheet VARCHAR(64) NOT NULL,
            InvoiceNo VARCHAR(32) NOT NULL,
            StockCode VARCHAR(64) NOT NULL,
            DescriptionText VARCHAR(512) NULL,
            Quantity INT NOT NULL,
            InvoiceDate DATETIME NOT NULL,
            UnitPrice DECIMAL(18, 4) NOT NULL,
            CustomerID VARCHAR(64) NULL,
            Country VARCHAR(128) NOT NULL,
            IsCancellation BOOLEAN NOT NULL,
            UnknownCustomer BOOLEAN NOT NULL,
            ZeroPrice BOOLEAN NOT NULL,
            KEY ix_invoice_date (InvoiceDate),
            KEY ix_country (Country),
            KEY ix_customer (CustomerID),
            KEY ix_stock_code (StockCode),
            KEY ix_product_cover (StockCode, DescriptionText)
        ) ENGINE=InnoDB
        """
    )
    ensure_product_refresh_schema(cursor, database, table)
    cursor.execute(f"SELECT COUNT(*) FROM `{table}`")
    existing_rows = int(cursor.fetchone()[0])
    if existing_rows and not replace:
        raise ValueError("raw_table_not_empty_use_replace")
    if existing_rows:
        cursor.execute(f"TRUNCATE TABLE `{table}`")


def insert_row(normalized: Mapping[str, Any], source_row_id: int, source_sheet: str) -> tuple[Any, ...]:
    invoice_date = normalized["InvoiceDate"]
    if not isinstance(invoice_date, datetime):
        invoice_date = datetime.fromisoformat(str(invoice_date))
    return (
        source_row_id,
        source_sheet,
        normalized["InvoiceNo"],
        normalized["StockCode"],
        normalized["Description"],
        normalized["Quantity"],
        invoice_date,
        normalized["UnitPrice"],
        normalized["CustomerID"],
        normalized["Country"],
        normalized["IsCancellation"],
        normalized["UnknownCustomer"],
        normalized["ZeroPrice"],
    )


def write_load_evidence(path: Path, payload: Mapping[str, Any]) -> None:
    path.mkdir(parents=True, exist_ok=True)
    (path / "mysql-load.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    lines = [
        "# MySQL raw load",
        "",
        f"- Status: **{payload['status']}**",
        f"- Workbook SHA-256: `{payload['workbook_sha256']}`",
        f"- Source rows observed: {payload['source_rows_observed']}",
        f"- Rows loaded: {payload['rows_loaded']}",
        f"- Rows rejected: {payload['rows_rejected']}",
        f"- Target: `{payload['database']}.{payload['table']}`",
        "",
        "## Rejection summary",
        "",
        "| Code | Count |",
        "|---|---:|",
    ]
    summary = payload["rejection_summary"]
    if summary:
        lines.extend(f"| {code} | {count} |" for code, count in sorted(summary.items()))
    else:
        lines.append("| none | 0 |")
    (path / "mysql-load.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def load_workbook_to_mysql(
    workbook_path: Path,
    values: Mapping[str, str],
    evidence_dir: Path,
    *,
    replace: bool,
) -> dict[str, Any]:
    try:
        import mysql.connector  # type: ignore[import-not-found]
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ModuleNotFoundError as error:
        raise RuntimeError("loader_dependency_missing") from error
    started_at = datetime.now(timezone.utc).isoformat()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    connection = mysql.connector.connect(
        host=values["MYSQL_HOST"],
        port=int(values["MYSQL_PORT"]),
        user=values["MYSQL_USER"],
        password=values["MYSQL_PASSWORD"],
        connection_timeout=20,
    )
    cursor = connection.cursor()
    database = values["MYSQL_DATABASE"]
    table = values["MYSQL_SOURCE_TABLE"]
    rejection_summary: dict[str, int] = {}
    source_rows_observed = 0
    rows_loaded = 0
    rows_rejected = 0
    batch: list[tuple[Any, ...]] = []
    insert_sql = f"""
        INSERT INTO `{database}`.`{table}`
        (SourceRowId, SourceSheet, InvoiceNo, StockCode, DescriptionText, Quantity, InvoiceDate,
         UnitPrice, CustomerID, Country, IsCancellation, UnknownCustomer, ZeroPrice)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    try:
        create_schema(cursor, database, table, replace=replace)
        source_row_id = 0
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if headers is None:
                continue
            header_names = [str(header).strip() if header is not None else "" for header in headers]
            for values_row in rows:
                source_row_id += 1
                source_rows_observed += 1
                raw_row = dict(zip(header_names, values_row, strict=True))
                try:
                    batch.append(insert_row(classify_row(raw_row), source_row_id, sheet.title))
                except ValueError as error:
                    rows_rejected += 1
                    code = str(error)
                    rejection_summary[code] = rejection_summary.get(code, 0) + 1
                    continue
                if len(batch) >= 5000:
                    cursor.executemany(insert_sql, batch)
                    connection.commit()
                    rows_loaded += len(batch)
                    batch.clear()
        if batch:
            cursor.executemany(insert_sql, batch)
            connection.commit()
            rows_loaded += len(batch)
    finally:
        workbook.close()
        cursor.close()
        connection.close()
    payload = {
        "name": "UCI Online Retail II MySQL raw load",
        "status": "passed",
        "started_at": started_at,
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "database": database,
        "table": table,
        "workbook_filename": workbook_path.name,
        "workbook_bytes": workbook_path.stat().st_size,
        "workbook_sha256": sha256(workbook_path),
        "source_rows_observed": source_rows_observed,
        "rows_loaded": rows_loaded,
        "rows_rejected": rows_rejected,
        "rejection_summary": rejection_summary,
    }
    write_load_evidence(evidence_dir, payload)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx-path", required=True, type=Path)
    parser.add_argument("--env-file", type=Path, default=PROJECT_ROOT / ".env")
    parser.add_argument("--evidence-dir", type=Path, default=PROJECT_ROOT / "evidence" / "data-quality")
    parser.add_argument("--replace", action="store_true")
    arguments = parser.parse_args()
    try:
        workbook_path = validate_workbook_path(arguments.xlsx_path, PROJECT_ROOT)
        payload = load_workbook_to_mysql(
            workbook_path,
            required_environment(arguments.env_file),
            arguments.evidence_dir,
            replace=arguments.replace,
        )
    except (OSError, RuntimeError, ValueError) as error:
        print(str(error))
        return 2
    print("mysql_load_passed")
    print(f"rows_loaded={payload['rows_loaded']}")
    print(f"rows_rejected={payload['rows_rejected']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
